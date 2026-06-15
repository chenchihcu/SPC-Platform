"""Drive ``build_pptx_report`` and ``export_diagnostic_summary_xlsx`` and verify
that the produced files are well-formed.

Both checks open the produced file with python-pptx / openpyxl to confirm that
not just the bytes exist, but the file can be parsed and contains at least one
slide / sheet.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from app.services.diagnostic_excel_exporter import export_diagnostic_summary_xlsx
from app.services.pptx_report_builder import build_pptx_report


def export_pptx(
    payload: dict[str, Any],
    spec: dict[str, Any],
    output_path: Path,
    chart_ids: list[str],
) -> tuple[bool, str]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    summary_data = (payload or {}).get("summary", {})
    try:
        ok, err = build_pptx_report(
            wo_master={},
            wo_spec=spec,
            summary_data=summary_data,
            diagnostics=[],
            output_path=str(output_path),
            analysis_payload=payload,
            report_context=None,
            template_type="engineering",
            chart_ids_to_export=list(chart_ids),
        )
    except BaseException as exc:
        return False, f"build_pptx_report raised: {type(exc).__name__}: {exc}"
    if not ok:
        return False, err or "build_pptx_report returned False"
    if not output_path.is_file() or output_path.stat().st_size < 1024:
        return False, f"pptx output missing or too small: {output_path}"
    try:
        from pptx import Presentation

        prs = Presentation(str(output_path))
        if len(prs.slides) < 1:
            return False, "pptx has 0 slides"
    except BaseException as exc:
        return False, f"pptx readback: {type(exc).__name__}: {exc}"
    return True, ""


def export_xlsx(payload: dict[str, Any], output_path: Path) -> tuple[bool, str]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        export_diagnostic_summary_xlsx(payload, str(output_path))
    except BaseException as exc:
        return False, f"export_xlsx raised: {type(exc).__name__}: {exc}"
    if not output_path.is_file() or output_path.stat().st_size < 512:
        return False, f"xlsx output missing or too small: {output_path}"
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(output_path))
        if not wb.sheetnames:
            return False, "xlsx has 0 sheets"
    except BaseException as exc:
        return False, f"xlsx readback: {type(exc).__name__}: {exc}"
    return True, ""
