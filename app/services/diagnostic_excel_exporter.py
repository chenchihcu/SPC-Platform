"""Excel export for DiagnosticPage summary payload."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore[import-untyped]

from app.analytics.dashboard_layers_display import (
    build_process_stat_report_sections,
    value_state_label_zh,
)
from app.services.diagnostic_evidence_matrix import build_readable_diagnostic_tabs
from app.ui.theme.tokens import (
    ACCENT_ERROR,
    ACCENT_SUCCESS,
    ACCENT_WARNING,
    ERROR_SURFACE_SUBTLE,
    EXCEL_BORDER,
    EXCEL_HEADER_BG,
    EXCEL_HEADER_TEXT,
    EXCEL_LABEL_BG,
    EXCEL_SECTION_BG,
    EXCEL_SECTION_TEXT,
    EXCEL_TEXT_PRIMARY,
    EXCEL_VALUE_BG,
    SUCCESS_SURFACE_SUBTLE,
    TEXT_SECONDARY,
    WARNING_SURFACE_SUBTLE,
)


_EXCEL_FONT_FAMILY = "Noto Sans TC"


def _xlsx_color(hex_color: str) -> str:
    return hex_color.strip().lstrip("#").upper()


def _solid_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=_xlsx_color(hex_color))


def _excel_font(*, size: int, bold: bool = False, color: str = EXCEL_TEXT_PRIMARY) -> Font:
    return Font(name=_EXCEL_FONT_FAMILY, size=size, bold=bold, color=_xlsx_color(color))


_HEADER_FILL = _solid_fill(EXCEL_HEADER_BG)
_SECTION_FILL = _solid_fill(EXCEL_SECTION_BG)
_LABEL_FILL = _solid_fill(EXCEL_LABEL_BG)
_VALUE_FILL = _solid_fill(EXCEL_VALUE_BG)
_HEADER_FONT = _excel_font(size=10, bold=True, color=EXCEL_HEADER_TEXT)
_SECTION_FONT = _excel_font(size=10, bold=True, color=EXCEL_SECTION_TEXT)
_LABEL_FONT = _excel_font(size=10, bold=True)
_VALUE_FONT = _excel_font(size=10)
_SOURCE_FONT = _excel_font(size=10, color=TEXT_SECONDARY)
_STATE_FILLS = {
    "good": _solid_fill(SUCCESS_SURFACE_SUBTLE),
    "warning": _solid_fill(WARNING_SURFACE_SUBTLE),
    "bad": _solid_fill(ERROR_SURFACE_SUBTLE),
    "neutral": _VALUE_FILL,
}
_STATE_FONTS = {
    "good": _excel_font(size=10, bold=True, color=ACCENT_SUCCESS),
    "warning": _excel_font(size=10, bold=True, color=ACCENT_WARNING),
    "bad": _excel_font(size=10, bold=True, color=ACCENT_ERROR),
    "neutral": _VALUE_FONT,
}
_BORDER_SIDE = Side(style="thin", color=_xlsx_color(EXCEL_BORDER))
_THIN_BORDER = Border(
    left=_BORDER_SIDE,
    right=_BORDER_SIDE,
    top=_BORDER_SIDE,
    bottom=_BORDER_SIDE,
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

_READABLE_SHEET_TITLES = {
    "overview": "總覽",
    "combination_matrix": "組合矩陣",
    "evidence_matrix": "證據矩陣",
    "correlation": "關聯判讀",
    "chart_linkage": "圖表連動",
    "actions": "對策建議",
    "data_context": "資料背景",
}


def _s(value: Any) -> str:
    return "—" if value in (None, "", []) else str(value)


def _section_row(ws, row: int, title: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = _SECTION_FILL
    cell.font = _SECTION_FONT
    cell.alignment = _LEFT
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = _THIN_BORDER


def _pair(ws, row: int, col: int, label: str, value: str) -> None:
    label_cell = ws.cell(row=row, column=col, value=label)
    value_cell = ws.cell(row=row, column=col + 1, value=value)
    label_cell.fill = _LABEL_FILL
    value_cell.fill = _VALUE_FILL
    label_cell.font = _LABEL_FONT
    value_cell.font = _VALUE_FONT
    label_cell.alignment = _LEFT
    value_cell.alignment = _LEFT
    label_cell.border = _THIN_BORDER
    value_cell.border = _THIN_BORDER


def _header_cells(ws, row: int, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill = _SECTION_FILL
        cell.font = _SECTION_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def _body_cells(ws, row: int, values: list[Any]) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=_s(value))
        cell.fill = _VALUE_FILL
        cell.font = _VALUE_FONT
        cell.alignment = _LEFT
        cell.border = _THIN_BORDER


def _report_header_cells(ws, row: int) -> None:
    headers = ["閱讀順序", "欄位", "數值", "嚴重性", "資料來源", "判讀用途"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill = _SECTION_FILL
        cell.font = _SECTION_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def _report_row_cells(
    ws,
    row: int,
    section_title: str,
    label: str,
    value: str,
    state: str,
    source: str,
    meaning: str,
) -> None:
    values = [
        section_title,
        label,
        value,
        value_state_label_zh(state),
        source,
        meaning,
    ]
    for col, raw_value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=_s(raw_value))
        cell.fill = _STATE_FILLS.get(state, _VALUE_FILL) if col in {3, 4} else _VALUE_FILL
        if col == 1:
            cell.fill = _LABEL_FILL
            cell.font = _SECTION_FONT
        elif col == 5:
            cell.font = _SOURCE_FONT
        elif col in {3, 4}:
            cell.font = _STATE_FONTS.get(state, _VALUE_FONT)
        else:
            cell.font = _VALUE_FONT
        cell.alignment = _WRAP if col in {2, 5, 6} else _LEFT
        cell.border = _THIN_BORDER


def _readable_body_cells(ws, row: int, values: list[Any]) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=_s(value))
        cell.fill = _VALUE_FILL
        cell.font = _VALUE_FONT
        cell.alignment = _WRAP
        cell.border = _THIN_BORDER


def _add_diagnostic_matrix_sheets(wb: Workbook, payload: dict[str, Any]) -> None:
    matrix = payload.get("diagnostic_evidence_matrix")
    if not isinstance(matrix, dict):
        return

    readable_tabs = build_readable_diagnostic_tabs(matrix)
    for tab_key, sheet_title in _READABLE_SHEET_TITLES.items():
        ws = wb.create_sheet(sheet_title)
        ws.sheet_view.showGridLines = False
        for col, width in {
            "A": 24,
            "B": 20,
            "C": 54,
            "D": 42,
            "E": 54,
            "F": 28,
        }.items():
            ws.column_dimensions[col].width = width
        _header_cells(ws, 1, ["項目", "判讀結果", "說明", "證據來源", "下一步", "資料來源"])
        rows = readable_tabs.get(tab_key, [])
        if not rows:
            rows = [
                {
                    "title": "目前無可判讀內容",
                    "result_zh": "資料不足/不可判讀",
                    "reason_zh": "此分頁目前沒有足夠資料形成白話判讀。",
                    "evidence_zh": "—",
                    "next_action_zh": "補齊資料或重跑分析後再查看。",
                    "source_zh": tab_key,
                }
            ]
        for row_idx, row in enumerate(rows[:300], start=2):
            if not isinstance(row, dict):
                continue
            _readable_body_cells(
                ws,
                row_idx,
                [
                    row.get("title"),
                    row.get("result_zh"),
                    row.get("reason_zh"),
                    row.get("evidence_zh"),
                    row.get("next_action_zh"),
                    row.get("source_zh"),
                ],
            )


def export_diagnostic_summary_xlsx(payload: dict[str, Any], output_path: str | Path) -> Path:
    """Export dashboard layer summary into a styled xlsx file."""
    summary = (payload or {}).get("summary", {})
    layers = summary.get("process", {}).get("dashboard_layers", {})
    if not isinstance(layers, dict):
        layers = {}

    out = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "診斷摘要"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22

    for col, width in {
        "A": 22,
        "B": 22,
        "C": 24,
        "D": 13,
        "E": 28,
        "F": 54,
    }.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "SMT SPI 製程統計分析報告"
    title_cell.fill = _HEADER_FILL
    title_cell.font = _HEADER_FONT
    title_cell.alignment = _CENTER

    ws.merge_cells("A2:F2")
    ts_cell = ws["A2"]
    ts_cell.value = f"產生時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ts_cell.fill = _LABEL_FILL
    ts_cell.font = _VALUE_FONT
    ts_cell.alignment = _LEFT

    _report_header_cells(ws, 4)
    row_idx = 5
    for section in build_process_stat_report_sections(layers):
        first_row = True
        for row_data in section["rows"]:
            _report_row_cells(
                ws,
                row_idx,
                section["title"] if first_row else "",
                row_data["label"],
                row_data["value"],
                row_data["state"],
                row_data["source"],
                row_data["meaning"],
            )
            ws.row_dimensions[row_idx].height = 30
            row_idx += 1
            first_row = False

    for row_num in range(1, row_idx):
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_num, column=col_idx)
            if cell.border != _THIN_BORDER and row_num in {1, 2}:
                cell.border = _THIN_BORDER

    _add_diagnostic_matrix_sheets(wb, payload)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
