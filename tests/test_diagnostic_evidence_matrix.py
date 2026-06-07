from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.services.diagnostic_evidence_matrix import (
    build_diagnostic_evidence_matrix,
    build_readable_diagnostic_tabs,
)
from app.services.diagnostic_excel_exporter import export_diagnostic_summary_xlsx
from app.services.report_process_narrative import build_process_diagnosis_report_payload


def _valid_payload(chart_type: str = "Generic") -> dict[str, Any]:
    return {
        "chart_type": chart_type,
        "data": {"values": [1, 2, 3, 4, 5]},
        "statistics": {"n": 5},
        "metadata": {"is_valid": True, "error": ""},
    }


def _cap(cpk: float = 1.6, cp: float = 1.7) -> dict[str, Any]:
    return {
        "chart_type": "Capability",
        "data": {},
        "statistics": {"cpk": cpk, "cp": cp, "mean": 100.0, "sigma_st": 2.0},
        "metadata": {"is_valid": True, "error": ""},
    }


def _spc(ooc_count: int = 0, n: int = 20) -> dict[str, Any]:
    return {
        "chart_type": "SPC",
        "data": {"out_of_control_indices": list(range(ooc_count)), "values": list(range(n))},
        "statistics": {"n": n},
        "metadata": {"is_valid": True, "error": ""},
    }


def _ooc(ooc_count: int = 0, n: int = 20) -> dict[str, Any]:
    ratio = ooc_count / n if n else None
    return {
        "chart_type": "OOCAnalysis",
        "data": {"ooc_count": ooc_count, "n": n, "ooc_ratio": ratio},
        "statistics": {"ooc_count": ooc_count, "n": n, "ooc_ratio": ratio},
        "metadata": {"is_valid": True, "error": ""},
    }


def _normality(is_normal: bool = True, p_value: float = 0.2) -> dict[str, Any]:
    return {
        "chart_type": "Normality",
        "data": {},
        "statistics": {"is_normal": is_normal, "p_value": p_value},
        "metadata": {"is_valid": True, "error": ""},
    }


def _drift(alarm: bool = False) -> dict[str, Any]:
    return {
        "chart_type": "DriftDetection",
        "data": {"trend_level": "Alarm Drift" if alarm else "Stable"},
        "statistics": {},
        "metadata": {"is_valid": True, "error": ""},
    }


def _cusum(ratio: float = 0.1) -> dict[str, Any]:
    return {
        "chart_type": "CUSUM",
        "data": {},
        "statistics": {"max_drift_ratio": ratio, "n": 20},
        "metadata": {"is_valid": True, "error": ""},
    }


def _corr(features: list[str], abs_corr: float = 0.2) -> dict[str, Any]:
    pairs: list[dict[str, Any]] = []
    for i, left in enumerate(features):
        for right in features[i + 1 :]:
            pairs.append(
                {
                    "pair": f"{left} vs {right}",
                    "left": left,
                    "right": right,
                    "corr": abs_corr,
                    "abs_corr": abs_corr,
                }
            )
    return {
        "chart_type": "CorrelationMatrix",
        "data": {"labels": features, "pairs_ranked": pairs},
        "statistics": {"n": 20, "pair_count": len(pairs), "strong_pair_count": int(abs_corr >= 0.7)},
        "metadata": {"is_valid": True, "error": ""},
    }


def _pass_fail(pass_rates: list[float] | None = None) -> dict[str, Any]:
    rates = pass_rates or [100.0, 100.0, 100.0]
    fails = [int(round((100.0 - r) / 100.0 * 20)) for r in rates]
    return {
        "chart_type": "PassFail",
        "data": {
            "labels": ["Volume", "Area", "Height"][: len(rates)],
            "pass_counts": [20 - f for f in fails],
            "fail_counts": fails,
            "pass_rates": rates,
            "denominator_n": [20 for _ in rates],
        },
        "statistics": {"n_total": 20},
        "metadata": {"is_valid": True, "error": ""},
    }


def _feature_bundle(feature: str, *, cpk: float = 1.6, cp: float = 1.7, ooc: int = 0, normal: bool = True, drift: bool = False) -> dict[str, Any]:
    cap = _cap(cpk=cpk, cp=cp)
    spc = _spc(ooc)
    return {
        "spc": spc,
        "xbar_r": _ooc(ooc),
        "cap": cap,
        "dist": _valid_payload("Histogram"),
        "box": _valid_payload("Boxplot"),
        "normality": _normality(normal, 0.004 if not normal else 0.2),
        "density": _valid_payload("Density"),
        "ewma": _valid_payload("EWMA"),
        "cusum": _cusum(1.2 if drift else 0.1),
        "run_chart": _valid_payload("RunChart"),
        "anova_parttype": _valid_payload("ANOVA"),
        "pattern_recognition": _valid_payload("PatternRecognition"),
        "ooc_analysis": _ooc(ooc),
        "shift_detection": _ooc(ooc),
        "drift_detection": _drift(drift),
        "outlier_analysis": _ooc(0),
        "subgroup": _valid_payload("Subgroup"),
        "repeated_offender": _valid_payload("RepeatedOffender"),
        "pareto": _valid_payload("Pareto"),
        "spatial": _valid_payload("Spatial"),
        "analysis_context": {"target_col": feature},
    }


def _payload(
    features: list[str],
    *,
    cpk: float = 1.6,
    cp: float = 1.7,
    ooc: int = 0,
    normal: bool = True,
    drift: bool = False,
    corr: float = 0.2,
    cluster_ratio: float = 0.0,
    pass_rates: list[float] | None = None,
) -> dict[str, Any]:
    parameters = {
        feature: _feature_bundle(feature, cpk=cpk, cp=cp, ooc=ooc, normal=normal, drift=drift)
        for feature in features
    }
    summary = {
        "per_measure": {
            feature: {"cap": {"statistics": {"cpk": cpk, "cp": cp}}, "n": 20}
            for feature in features
        },
        "relation": {
            "corr_vol_area": corr,
            "corr_vol_height": corr,
            "corr_area_height": corr,
        },
        "process": {
            "verdict": "可接受" if cpk >= 1.33 else "待改善",
            "dashboard_layers": {
                "layer_1_alarm": {
                    "issue_type_display_zh": "—",
                    "max_drift_ratio": 1.2 if drift else 0.0,
                    "ooc_count": ooc,
                },
                "layer_2_kpi": {"yield_pct": 99.0, "dpmo": 100.0, "sigma_level": 4.0},
                "layer_3_info": {"sample_size": 20, "range": 5, "driver_feature": features[0]},
                "layer_4_defect_structure": {
                    "cluster_ratio": cluster_ratio,
                    "top_oos_refdes": [{"id": "R1", "oos_count": 3}] if cluster_ratio > 0 else [],
                },
                "layer_5_spec_analysis": {
                    "cpk": cpk,
                    "cp": cp,
                    "oos_rate": 0.02 if cpk < 1.0 else 0.0,
                    "std_spec_ratio": 0.1,
                },
                "layer_6_product_context": {"product_name": "PN"},
                "layer_7_engineering_info": {"sample_size": 20, "mean": 100.0, "std": 2.0},
                "layer_8_diagnosis": {"priority": "low", "issue_type_display_zh": "—"},
            },
        },
    }
    payload: dict[str, Any] = {
        "selected_features": features,
        "summary": summary,
        "parameters": parameters,
        "cap": parameters[features[0]]["cap"],
        "spc": parameters[features[0]]["spc"],
        "normality": parameters[features[0]]["normality"],
        "cusum": parameters[features[0]]["cusum"],
        "drift_detection": parameters[features[0]]["drift_detection"],
        "ooc_analysis": parameters[features[0]]["ooc_analysis"],
        "spatial": parameters[features[0]]["spatial"],
        "pareto": parameters[features[0]]["pareto"],
        "correlation_matrix": _corr(features, corr) if len(features) >= 2 else None,
        "correlation_heatmap": _corr(features, corr) if len(features) >= 2 else None,
        "pass_fail_matrix": _pass_fail(pass_rates) if len(features) == 3 else None,
        "anomaly_3f": _valid_payload("Anomaly3F") if len(features) == 3 else None,
        "consistency_3f": _valid_payload("Consistency3F") if len(features) == 3 else None,
        "parallel_coord": _valid_payload("ParallelCoord") if len(features) == 3 else None,
        "triple_parameters": {
            "pass_fail_matrix": _pass_fail(pass_rates),
            "anomaly_3f": _valid_payload("Anomaly3F"),
            "consistency_3f": _valid_payload("Consistency3F"),
            "parallel_coord": _valid_payload("ParallelCoord"),
        }
        if len(features) == 3
        else {},
    }
    return payload


def test_triple_feature_combination_expands_by_chart_arity() -> None:
    payload = _payload(["Volume", "Area", "Height"])
    matrix = build_diagnostic_evidence_matrix(payload)

    combo = matrix["combination_summary"]
    assert combo["single_feature_candidate_count"] > 0
    assert combo["dual_feature_candidate_count"] > 0
    assert combo["triple_feature_candidate_count"] > 0
    assert combo["candidate_count"] == (
        combo["single_feature_candidate_count"]
        + combo["dual_feature_candidate_count"]
        + combo["triple_feature_candidate_count"]
    )
    assert any(
        c["chart_id"] == "correlation_matrix"
        and c["feature_set"] == ["Volume", "Area"]
        and c["availability"] == "analyzed"
        for c in matrix["candidates"]
    )
    assert any(
        c["chart_id"] == "pass_fail_matrix" and c["availability"] == "analyzed"
        for c in matrix["candidates"]
    )


def test_single_feature_marks_dual_and_triple_charts_not_applicable() -> None:
    matrix = build_diagnostic_evidence_matrix(_payload(["Volume"]))

    assert any(c["chart_id"] == "correlation_matrix" and c["availability"] == "not-applicable" for c in matrix["candidates"])
    assert any(c["chart_id"] == "pass_fail_matrix" and c["availability"] == "not-applicable" for c in matrix["candidates"])
    assert matrix["coverage"]["applicable_candidate_count"] < matrix["coverage"]["candidate_count"]


def test_readable_tabs_keep_internal_refute_but_use_plain_language() -> None:
    matrix = build_diagnostic_evidence_matrix(_payload(["Volume"], cpk=1.6, cp=1.8))

    assert any(c["evidence_state"] == "refute" for c in matrix["candidates"])

    readable = build_readable_diagnostic_tabs(matrix)
    joined = "\n".join(
        str(value)
        for rows in readable.values()
        for row in rows
        for value in row.values()
    )
    assert "不支持此假設" in joined
    assert "反證" not in joined
    assert all(
        {"title", "result_zh", "reason_zh", "evidence_zh", "next_action_zh", "source_zh"} <= set(row)
        for rows in readable.values()
        for row in rows
    )


def test_cpk_alone_does_not_become_high_confidence_root_cause() -> None:
    matrix = build_diagnostic_evidence_matrix(_payload(["Volume"], cpk=0.82, cp=1.8))

    conf = matrix["summary"]["confidence"]
    assert conf["level"] != "high"
    assert conf["support_family_count"] == 1
    assert "不足以單獨定根因" in matrix["summary"]["verdict_zh"]
    assert matrix["summary"]["conflicts"]


def test_capability_plus_pass_fail_matrix_sets_capability_pattern() -> None:
    payload = _payload(["Volume", "Area", "Height"], cpk=0.82, cp=1.8, pass_rates=[90.0, 100.0, 100.0])
    matrix = build_diagnostic_evidence_matrix(payload)

    assert matrix["summary"]["pattern_label"] == "規格/能力主導異常"
    assert any(item["chart_id"] == "pass_fail_matrix" for item in matrix["summary"]["top_evidence"])


def test_drift_and_nonnormal_use_multi_signal_correlation_rules() -> None:
    payload = _payload(["Volume"], normal=False, drift=True)
    matrix = build_diagnostic_evidence_matrix(payload)

    assert matrix["summary"]["pattern_label"] == "系統性漂移 + 分布異常"


def test_filter_context_changes_scope_text() -> None:
    matrix = build_diagnostic_evidence_matrix(
        _payload(["Volume", "Area"]),
        filter_context={"batch": "首件", "refdes": "R1", "part_type": "R0402"},
    )

    scope = matrix["filter_scope"]["scope_zh"]
    assert "元件=R1" in scope
    assert "類別=R0402" in scope
    assert "範圍=首件" in scope


def test_report_narrative_and_excel_export_include_matrix(tmp_path: Path) -> None:
    payload = _payload(["Volume", "Area", "Height"], cpk=0.82, cp=1.8, pass_rates=[90.0, 100.0, 100.0])
    payload["diagnostic_evidence_matrix"] = build_diagnostic_evidence_matrix(payload)

    pdr = build_process_diagnosis_report_payload(payload)
    assert pdr["schema_version"] == "1.1.0"
    assert pdr["C_evidence"]["combination_coverage_zh"]
    assert pdr["C_evidence"]["top_evidence"]
    assert pdr["C_evidence"]["readable_evidence"]

    out = export_diagnostic_summary_xlsx(payload, tmp_path / "diag.xlsx")
    wb = load_workbook(out)
    summary_ws = wb["診斷摘要"]
    assert summary_ws["A1"].value == "SMT SPI 製程統計分析報告"
    assert [summary_ws.cell(row=4, column=col).value for col in range(1, 7)] == [
        "閱讀順序",
        "欄位",
        "數值",
        "嚴重性",
        "資料來源",
        "判讀用途",
    ]
    summary_rows = {
        str(summary_ws.cell(row=row, column=2).value): row
        for row in range(5, summary_ws.max_row + 1)
        if summary_ws.cell(row=row, column=2).value
    }
    cpk_row = summary_rows["Cpk (主特徵)"]
    oos_row = summary_rows["OOS 比率 (規格)"]
    usl_row = summary_rows["USL"]
    assert summary_ws.cell(cpk_row, 4).value == "需處置"
    assert summary_ws.cell(oos_row, 4).value == "需處置"
    assert summary_ws.cell(cpk_row, 5).value == "layer_5_spec_analysis"
    assert summary_ws.cell(oos_row, 5).value == "layer_5_spec_analysis"
    assert (
        summary_ws.cell(cpk_row, 3).fill.fgColor.rgb
        != summary_ws.cell(usl_row, 3).fill.fgColor.rgb
    )

    expected_sheets = {"總覽", "組合矩陣", "證據矩陣", "關聯判讀", "圖表連動", "對策建議", "資料背景"}
    assert expected_sheets <= set(wb.sheetnames)
    for sheet_name in expected_sheets:
        ws = wb[sheet_name]
        headers = [ws.cell(row=1, column=col).value for col in range(1, 7)]
        assert headers == ["項目", "判讀結果", "說明", "證據來源", "下一步", "資料來源"]
    visible_text = "\n".join(
        str(cell.value)
        for sheet_name in expected_sheets
        for row in wb[sheet_name].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "refute" not in visible_text
    assert "反證" not in visible_text
